import time
import json
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException, NoSuchElementException, WebDriverException
import telebot
import openpyxl
from openpyxl import Workbook, load_workbook
import os
import traceback
import urllib.parse
import time
import signal
import logging
from datetime import datetime

# Global flag to control the main loop
running = True

# Set up logging
logging.basicConfig(filename='carousell_scraper.log', level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s')


def signal_handler(signum, frame):
    global running
    print("\nReceived signal to stop. Finishing current iteration and then stopping...")
    running = False


# Register the signal handler
signal.signal(signal.SIGINT, signal_handler)

step_counter = 0


def log(message):
    global step_counter
    step_counter += 1
    log_message = f"[Step {step_counter}]: {message}"
    print(log_message)
    logging.info(log_message)


# Load configuration


def load_config():
    log("Loading configuration...")
    try:
        with open('config.json', 'r') as f:
            config = json.load(f)
            log("Configuration loaded successfully.")
            return config
    except FileNotFoundError:
        log("Configuration file not found. Please create a config.json file.")
        sys.exit(1)
    except json.JSONDecodeError:
        log("Error decoding JSON from config file. Please check the format.")
        sys.exit(1)


config = load_config()
# Default to 48 if not specified
MAX_LISTINGS = config.get('MAX_LISTINGS_TO_SCRAPE', 48)

# Initialize the Telegram bot
log("Initializing Telegram bot...")
try:
    bot = telebot.TeleBot(config['TELEGRAM_BOT_TOKEN'])
    log("Telegram bot initialized successfully.")
except Exception as e:
    log(f"Error initializing Telegram bot: {str(e)}")
    bot = None


def send_telegram_message(message):
    if bot:
        try:
            log(f"Sending Telegram message: {message}")
            bot.send_message(config['TELEGRAM_CHAT_ID'], message)
            log("Telegram message sent successfully.")
        except Exception as e:
            log(f"Error sending Telegram message: {str(e)}")
    else:
        log("Telegram bot not initialized. Message not sent.")


def load_existing_ids(excel_path):
    log(f"Loading existing IDs from Excel file: {excel_path}")
    try:
        if os.path.exists(excel_path):
            log("Excel file found. Loading existing IDs...")
            workbook = load_workbook(excel_path)
            sheet = workbook.active
            existing_ids = {sheet.cell(
                row=row, column=1).value for row in range(2, sheet.max_row + 1)}
            log(f"Loaded {len(existing_ids)} existing IDs.")
        else:
            log("Excel file not found. Creating a new workbook...")
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Listing ID", "Href", "Seller", "Time Posted",
                         "Product Details", "Price", "Condition", "Image URL"])
            existing_ids = set()
            workbook.save(excel_path)
            log("New Excel file created and saved.")
        return existing_ids, workbook, sheet
    except Exception as e:
        log(f"Error handling Excel file: {str(e)}")
        return set(), Workbook(), Workbook().active


def save_to_excel(data, excel_path, sheet):
    try:
        log(f"Saving {len(data)} new listings to Excel...")
        for listing in data:
            sheet.append(listing)
        sheet.parent.save(excel_path)
        log("Excel file updated and saved successfully.")
    except PermissionError:
        log(f"Error: Permission denied when trying to save to {excel_path}")
        log("Please ensure the file is not open in another program and that you have write permissions.")
        alternative_path = os.path.join(os.path.expanduser(
            "~"), "Desktop", "carousell_listings.xlsx")
        log(f"Attempting to save to alternative location: {alternative_path}")
        try:
            sheet.parent.save(alternative_path)
            log(
                f"Successfully saved to alternative location: {alternative_path}")
        except Exception as e:
            log(f"Error saving to alternative location: {str(e)}")
    except Exception as e:
        log(f"Error saving to Excel: {str(e)}")
        log(f"Current working directory: {os.getcwd()}")
        log(f"Full path of excel file: {os.path.abspath(excel_path)}")
        log(f"File exists: {os.path.exists(excel_path)}")
        if os.path.exists(excel_path):
            log(f"File is writable: {os.access(excel_path, os.W_OK)}")
            log(f"Directory is writable: {os.access(os.path.dirname(excel_path), os.W_OK)}")


def build_url(search_item):
    if 'full_url' in search_item:
        return search_item['full_url']

    base_url = config['BASE_URL']
    params = {
        "search": search_item['query'],
        "price_start": search_item['price_start'],
        "price_end": search_item['price_end'],
        "sort_by": search_item['sort_by'],
    }

    if 'tab' in search_item and search_item['tab'] != 'all':
        params['tab'] = search_item['tab']

    category = search_item['category']
    url = f"{base_url}categories/{category}/"

    # Encode the query parameters
    encoded_params = urllib.parse.urlencode(
        {k: v for k, v in params.items() if v is not None},
        quote_via=urllib.parse.quote
    )

    if encoded_params:
        url += f"?{encoded_params}"

    return url


def find_element_with_fallback(card, selectors):
    for selector_type, selector in selectors:
        try:
            if selector_type == By.XPATH:
                element = card.find_element(By.XPATH, selector)
            elif selector_type == By.CSS_SELECTOR:
                element = card.find_element(By.CSS_SELECTOR, selector)
            if element:
                return element
        except NoSuchElementException:
            continue
    return None


def extract_text_content(element):
    return element.text.strip() if element else 'Not found'


def find_element_by_text(driver, text):
    try:
        return driver.find_element(By.XPATH, f"//p[contains(text(), '{text}')]")
    except NoSuchElementException:
        return None


def find_element_by_text_in_card(card, text):
    try:
        return card.find_element(By.XPATH, f".//p[contains(text(), '{text}')]")
    except NoSuchElementException:
        return None


def find_image(card):
    selectors = [
        ".//img[contains(@class, 'D_QJ')]",
        ".//img[contains(@class, 'D_SC')]",
        ".//img"  # fallback to any image
    ]
    for selector in selectors:
        try:
            return card.find_element(By.XPATH, selector)
        except NoSuchElementException:
            continue
    return None


def analyze_listing_card(card):
    title_selectors = [
        # This class seems to be used for titles
        (By.XPATH, ".//p[contains(@class, 'D_lO')]"),
        (By.XPATH,
         ".//a[contains(@class, 'D_ml')]//p[contains(@style, 'max-line')]"),
        (By.XPATH, ".//p[contains(@class, 'D_mb') and not(@data-testid)]")
    ]
    price_selectors = [
        (By.XPATH, ".//p[contains(@class, 'D_ma')]"),
        (By.XPATH, ".//p[contains(@class, 'D_mc')]"),
        (By.XPATH, ".//p[contains(text(), 'S$')]")
    ]
    time_selectors = [
        (By.XPATH, ".//p[contains(@class, 'D_ow')]"),
        (By.XPATH, ".//p[contains(@class, 'D_pc')]"),
        (By.XPATH, ".//p[contains(text(), 'ago')]")
    ]

    title = find_element_with_fallback(card, title_selectors)
    # Add a debug log to check what's being extracted
    log(f"Debug - Raw title text: {title.text if title else 'Not found'}")

    price = find_element_with_fallback(card, price_selectors)
    seller_name = card.find_element(
        By.XPATH, ".//p[@data-testid='listing-card-text-seller-name']")
    time = find_element_with_fallback(card, time_selectors)

    condition_types = ['Brand new', 'Like new',
                       'Lightly used', 'Well used', 'Heavily used']
    condition = None
    for condition_type in condition_types:
        condition_element = find_element_by_text_in_card(card, condition_type)
        if condition_element:
            condition = condition_element
            break

    listing_id = card.get_attribute('data-testid').replace('listing-card-', '')
    image = find_image(card)

    return {
        'id': listing_id,
        'title': extract_text_content(title),
        'price': extract_text_content(price),
        'seller_name': extract_text_content(seller_name),
        'time': extract_text_content(time),
        'condition': extract_text_content(condition),
        'image_url': image.get_attribute('src') if image else 'Not found',
        'href': f"https://www.carousell.sg/p/{listing_id}"
    }


def check_carousell_listings():
    log("Starting to check Carousell listings...")
    chrome_options = Options()
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")

    try:
        log("Initializing Chrome driver...")
        driver = webdriver.Chrome(service=Service(
            ChromeDriverManager().install()), options=chrome_options)

        excel_path = "carousell_listings.xlsx"
        log(f"Loading existing IDs from Excel: {excel_path}")
        existing_ids, workbook, sheet = load_existing_ids(excel_path)
        new_listings = []

        for search_item in config['SEARCH_ITEMS']:
            url = build_url(search_item)
            log(f"Navigating to URL: {url}")
            driver.get(url)

            log(f"Waiting for page to load (timeout: 45 seconds)...")

            try:
                WebDriverWait(driver, 45).until(
                    EC.presence_of_element_located((By.TAG_NAME, "body"))
                )
                log("Initial page content loaded.")

                WebDriverWait(driver, 45).until(
                    EC.presence_of_element_located(
                        (By.XPATH, "//div[contains(@class, 'browse-listings')]"))
                )
                log("Listing cards container found.")

            except TimeoutException:
                log("Timeout occurred while waiting for the page to load.")
                log("Current page source:")
                log(driver.page_source)
                continue

            listing_cards = driver.find_elements(
                By.XPATH, "//div[starts-with(@data-testid, 'listing-card-')]")
            log(f"Found {len(listing_cards)} listing cards.")

            if not listing_cards:
                log("No listing cards found. Possible page structure change.")
                log("Current page source:")
                log(driver.page_source)
                continue

            for index, card in enumerate(listing_cards[:MAX_LISTINGS]):
                log(f"Processing listing {index + 1}...")
                try:
                    listing_data = analyze_listing_card(card)
                    log(
                        f"Extracted details - Title: {listing_data['title']}, Price: {listing_data['price']}, Condition: {listing_data['condition']}")

                    if listing_data['id'] in existing_ids:
                        log("Listing already exists in database. Skipping.")
                        continue

                    price_value = float(listing_data['price'].replace(
                        "S$", "").replace(",", ""))

                    # Add these debug log statements here
                    log(
                        f"Debug - search_item['query']: {search_item.get('query', 'N/A')}")
                    log(
                        f"Debug - listing_data['title']: {listing_data['title']}")
                    log(
                        f"Debug - search_item['price_start']: {search_item.get('price_start', 'N/A')}")
                    log(
                        f"Debug - search_item['price_end']: {search_item.get('price_end', 'N/A')}")
                    log(f"Debug - price_value: {price_value}")

                    # Check if it's a full_url search_item or a regular one
                    if 'full_url' in search_item:
                        # For full_url items, we don't have specific criteria, so we add all listings
                        log("Full URL search item. Adding to new listings.")
                        new_listing = True
                    elif (search_item['query'].lower() in listing_data['title'].lower() and
                          (search_item['price_start'] is None or price_value >= search_item['price_start']) and
                            (search_item['price_end'] is None or price_value <= search_item['price_end'])):
                        log("Listing matches criteria. Adding to new listings.")
                        new_listing = True
                    else:
                        log("Listing does not match criteria. Skipping.")
                        new_listing = False

                    if new_listing:
                        new_listings.append([
                            listing_data['id'], listing_data['href'], listing_data['seller_name'],
                            listing_data['time'], listing_data['title'], listing_data['price'],
                            listing_data['condition'], listing_data['image_url']
                        ])
                        message = f"New listing found!\nTitle: {listing_data['title']}\nPrice: {listing_data['price']}\nCondition: {listing_data['condition']}\nSeller: {listing_data['seller_name']}\nPosted: {listing_data['time']}\nLink: {listing_data['href']}"
                        send_telegram_message(message)
                        log("Telegram message sent for new listing.")

                except Exception as e:
                    log(f"Error processing listing: {str(e)}")
                    log(f"Listing HTML: {card.get_attribute('outerHTML')}")

        if new_listings:
            log(f"Saving {len(new_listings)} new listings to Excel...")
            save_to_excel(new_listings, excel_path, sheet)
        else:
            log("No new listings found.")

    except Exception as e:
        log(f"Unexpected error: {str(e)}")
        log(traceback.format_exc())
        return False
    finally:
        driver.quit()
        log("Chrome driver closed.")
    return True


def main():
    global running
    log("Starting main loop. Press Ctrl+C to stop safely.")
    while running:
        try:
            log("Checking Carousell listings...")
            success = check_carousell_listings()
            if success:
                log("Successfully checked listings.")
            else:
                log("Failed to check listings.")

            if running:
                log("Waiting for 1 hour before next check...")
                for _ in range(3600):  # 3600 seconds = 1 hour
                    if not running:
                        break
                    time.sleep(1)
        except Exception as e:
            log(f"Unexpected error in main loop: {str(e)}")
            log("Continuing to next iteration...")

    log("Script stopped gracefully.")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        log(f"Critical error: {str(e)}")
    finally:
        log("Script execution completed.")
